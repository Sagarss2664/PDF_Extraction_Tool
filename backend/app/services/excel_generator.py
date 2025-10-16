# excel_generator.py
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging
from typing import Dict, Any, List, Tuple
import json
import os
from datetime import datetime
import re

logger = logging.getLogger(__name__)

class ExcelGenerator:
    def __init__(self):
        self.colors = {
            'header_blue': '2E86AB',
            'light_gray': 'F8F9FA', 
            'white': 'FFFFFF',
            'dark_gray': '6C757D',
            'warning_red': 'DC3545',
            'success_green': '28A745',
            'accent_green': '20C997',
            'accent_orange': 'FD7E14'
        }
        
        # Enhanced template configurations with specific field mappings
        self.template_configs = {
            1: {
                'name': 'Private Equity Fund Detailed Template',
                'color': 'header_blue',
                'sheets': {
                    'Fund_and_Investment_Vehicle_Information': 'Fund and Investment Vehicle Information',
                    'Fund_Manager': 'Fund Manager',
                    'Fund_Investment_Vehicle_Financial_Position': 'Fund Investment Vehicle Financial Position',
                    'LP_Investor_Cashflows': 'LP Investor Cashflows',
                    'Fund_Companies': 'Fund Companies',
                    'Initial_Investments': 'Initial Investments',
                    'Company_Investment_Positions': 'Company Investment Positions',
                    'Company_Valuation': 'Company Valuation',
                    'Company_Financials': 'Company Financials',
                    'Investment_History': 'Investment History',
                    'Reference_Values': 'Reference Values'
                }
            },
            2: {
                'name': 'Portfolio Summary Template',
                'color': 'accent_green',
                'sheets': {
                    'Executive_Portfolio_Summary': 'Executive Portfolio Summary',
                    'Schedule_of_Investments': 'Schedule of Investments',
                    'Statement_of_Operations': 'Statement of Operations',
                    'Statements_of_Cashflows': 'Statements of Cashflows',
                    'PCAP_Statements': 'PCAP Statements',
                    'Portfolio_Companies_Profile': 'Portfolio Companies Profile',
                    'Portfolio_Companies_Financials': 'Portfolio Companies Financials',
                    'FootNotes': 'FootNotes',
                    'Reference_Values': 'Reference Values'
                }
            }
        }
    
    def generate_excel(self, structured_data: Dict[str, Any], template_id: int, output_path: str, pdf_file_path: str = None) -> str:
        """
        Generate Excel file from structured data with PDF name as Excel file name
        
        Args:
            structured_data: Extracted data from LLM processor
            template_id: Template ID (1 or 2)
            output_path: Can be either full file path or directory path
            pdf_file_path: Path to the source PDF file (optional)
            
        Returns:
            str: Path to the generated Excel file
        """
        try:
            logger.info("🚀 STARTING EXCEL GENERATION 🚀")
            
            # Determine the actual output path
            excel_file_path = self._determine_output_path(output_path, pdf_file_path, template_id)
            
            logger.info(f"PDF File: {pdf_file_path}")
            logger.info(f"Output Excel: {excel_file_path}")
            logger.info(f"Template ID: {template_id}")
            
            # Debug: Log the complete data structure
            logger.info("=== STRUCTURED DATA RECEIVED ===")
            logger.info(f"Data type: {type(structured_data)}")
            logger.info(f"Data keys: {list(structured_data.keys()) if structured_data else 'EMPTY'}")

            # Create workbook
            workbook = Workbook()
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])

            # Get PDF file name for display
            pdf_file_name = os.path.basename(pdf_file_path) if pdf_file_path else "Unknown PDF"
            
            # Create overview sheet with template name and PDF name
            self._create_overview_sheet(workbook, structured_data, template_id, pdf_file_name)
            
            # Create data sheets based on template with enhanced mapping
            if template_id == 1:
                success = self._create_template_1_sheets(workbook, structured_data)
            elif template_id == 2:
                success = self._create_template_2_sheets(workbook, structured_data)
            else:
                success = self._create_generic_sheets(workbook, structured_data)

            # Save the workbook
            os.makedirs(os.path.dirname(excel_file_path) if os.path.dirname(excel_file_path) else '.', exist_ok=True)
            workbook.save(excel_file_path)
            
            logger.info(f"✅ Excel file successfully generated: {excel_file_path}")
            logger.info(f"📊 Sheets created: {workbook.sheetnames}")
            
            return excel_file_path
            
        except Exception as e:
            logger.error(f"❌ Error generating Excel file: {str(e)}")
            import traceback
            logger.error(f"Stack trace: {traceback.format_exc()}")
            raise Exception(f"Failed to generate Excel file: {str(e)}")

    def _determine_output_path(self, output_path: str, pdf_file_path: str, template_id: int) -> str:
        """
        Determine the final output path based on input parameters
        
        Args:
            output_path: Can be directory or full file path
            pdf_file_path: Source PDF file path
            template_id: Template ID for suffix
            
        Returns:
            str: Full path to Excel file
        """
        # If output_path is a directory, generate file name from PDF
        if os.path.isdir(output_path) or (output_path and not output_path.endswith('.xlsx')):
            if pdf_file_path:
                return self._generate_excel_file_path(pdf_file_path, output_path, template_id)
            else:
                # Fallback: use timestamp-based name
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                return os.path.join(output_path, f"extracted_data_{timestamp}.xlsx")
        else:
            # output_path is already a full file path
            return output_path

    def _generate_excel_file_path(self, pdf_file_path: str, output_directory: str, template_id: int) -> str:
        """
        Generate Excel file path based on PDF name and template
        
        Args:
            pdf_file_path: Path to the source PDF file
            output_directory: Directory to save Excel file
            template_id: Template ID for suffix
            
        Returns:
            str: Full path to Excel file
        """
        # Get PDF file name without extension
        pdf_name = os.path.splitext(os.path.basename(pdf_file_path))[0]
        
        # Clean the file name (remove invalid characters)
        clean_name = self._clean_file_name(pdf_name)
        
        # Add template suffix if needed
        if template_id == 1:
            file_suffix = "_Private_Equity"
        elif template_id == 2:
            file_suffix = "_Portfolio_Summary"
        else:
            file_suffix = "_Extracted"
        
        # Construct full path
        excel_file_name = f"{clean_name}{file_suffix}.xlsx"
        excel_file_path = os.path.join(output_directory, excel_file_name)
        
        logger.info(f"Generated Excel file path: {excel_file_path}")
        return excel_file_path

    def _clean_file_name(self, file_name: str) -> str:
        """
        Clean file name by removing invalid characters
        
        Args:
            file_name: Original file name
            
        Returns:
            str: Cleaned file name
        """
        # Remove invalid characters for file names
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            file_name = file_name.replace(char, '_')
        
        # Replace multiple spaces with single underscore
        file_name = re.sub(r'\s+', '_', file_name)
        
        # Remove leading/trailing spaces and underscores
        file_name = file_name.strip(' _')
        
        # Ensure name is not too long
        if len(file_name) > 100:
            file_name = file_name[:100]
        
        return file_name

    def _create_overview_sheet(self, workbook: Workbook, data: Dict[str, Any], template_id: int, pdf_file_name: str = None):
        """Create an overview sheet showing template and PDF information"""
        # Get template name
        template_name = self.template_configs.get(template_id, {}).get('name', f'Template {template_id}')
        
        # Create sheet name based on template and PDF
        if pdf_file_name:
            base_name = os.path.splitext(pdf_file_name)[0]
            sheet_name = f"{base_name} - {template_name}"[:31]  # Truncate to 31 chars
        else:
            sheet_name = template_name[:31]
        
        sheet = workbook.create_sheet(sheet_name)
        row_idx = 1
        
        # Main Title with Template Name and PDF
        title_text = f"{template_name}"
        if pdf_file_name:
            title_text += f" - {pdf_file_name}"
        
        title_cell = sheet.cell(row=row_idx, column=1, value=title_text)
        title_cell.font = Font(size=16, bold=True, color=self.colors['header_blue'])
        row_idx += 2
        
        # Extraction Information
        sheet.cell(row=row_idx, column=1, value="Extraction Information").font = Font(size=12, bold=True)
        row_idx += 1
        
        info_data = [
            ("Template ID:", template_id),
            ("Template Name:", template_name),
            ("PDF File:", pdf_file_name or "Not specified"),
            ("Extraction Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Data Type:", str(type(data))),
        ]
        
        for label, value in info_data:
            sheet.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            sheet.cell(row=row_idx, column=2, value=value)
            row_idx += 1
        
        row_idx += 1
        
        if not data:
            sheet.cell(row=row_idx, column=1, value="❌ NO DATA RECEIVED - Check LLM processing").font = Font(color=self.colors['warning_red'], bold=True)
            return
        
        # Data Summary
        sheet.cell(row=row_idx, column=1, value="Data Summary").font = Font(size=12, bold=True)
        row_idx += 1
        
        sheet.cell(row=row_idx, column=1, value="Available Sections:").font = Font(bold=True)
        available_sections = [key for key in data.keys() if not key.startswith('_') and self._has_meaningful_data(data.get(key))]
        sheet.cell(row=row_idx, column=2, value=", ".join(available_sections) if available_sections else "None")
        row_idx += 1
        
        sheet.cell(row=row_idx, column=1, value="Total Data Points:").font = Font(bold=True)
        sheet.cell(row=row_idx, column=2, value=self._count_data_points(data))
        row_idx += 2
        
        # Show all data in a structured way (only if meaningful data exists)
        if self._has_meaningful_data(data):
            sheet.cell(row=row_idx, column=1, value="Extracted Data Structure").font = Font(size=12, bold=True)
            row_idx += 2
            
            self._add_structured_overview(sheet, data, row_idx, 1)
        
        # Auto-adjust columns
        self._auto_adjust_columns(sheet)

    def _add_structured_overview(self, sheet, data: Any, start_row: int, start_col: int, level: int = 0):
        """Recursively add structured overview data to sheet"""
        current_row = start_row
        
        if data is None:
            sheet.cell(row=current_row, column=start_col, value="NULL")
            return current_row + 1
        elif isinstance(data, (str, int, float, bool)):
            sheet.cell(row=current_row, column=start_col, value=f"{'  ' * level}{str(data)}")
            return current_row + 1
        elif isinstance(data, dict):
            for key, value in data.items():
                if key.startswith('_'):  # Skip metadata in overview
                    continue
                    
                # Key with indentation
                key_display = f"{'  ' * level}📁 {key}:"
                sheet.cell(row=current_row, column=start_col, value=key_display).font = Font(bold=True)
                
                if isinstance(value, (dict, list)) and value and self._has_meaningful_data(value):
                    current_row += 1
                    current_row = self._add_structured_overview(sheet, value, current_row, start_col + 1, level + 1)
                else:
                    display_value = self._format_overview_value(value)
                    sheet.cell(row=current_row, column=start_col + 1, value=display_value)
                    current_row += 1
            return current_row
        elif isinstance(data, list):
            for i, item in enumerate(data):
                if i >= 5:  # Limit to first 5 items in overview
                    sheet.cell(row=current_row, column=start_col, value=f"{'  ' * level}... and {len(data) - 5} more items")
                    current_row += 1
                    break
                    
                sheet.cell(row=current_row, column=start_col, value=f"{'  ' * level}📄 [{i}]:")
                current_row += 1
                current_row = self._add_structured_overview(sheet, item, current_row, start_col + 1, level + 1)
            return current_row
        else:
            sheet.cell(row=current_row, column=start_col, value=f"{'  ' * level}UNKNOWN: {str(data)}")
            return current_row + 1

    def _format_overview_value(self, value: Any) -> str:
        """Format values for overview display"""
        if value is None:
            return "N/A"
        elif isinstance(value, (str, int, float, bool)):
            return str(value)
        elif isinstance(value, list):
            if not value:
                return "Empty list"
            return f"List with {len(value)} items"
        elif isinstance(value, dict):
            return f"Object with {len(value)} fields"
        else:
            return str(value)

    def _create_template_1_sheets(self, workbook: Workbook, data: Dict[str, Any]) -> bool:
        """Create sheets for Template 1 - Private Equity Fund Detailed Template"""
        logger.info("Creating Template 1 sheets...")
        
        template_config = self.template_configs[1]
        data_found = False
        
        # Process each expected section for Template 1
        for data_key, sheet_name in template_config['sheets'].items():
            section_data = data.get(data_key)
            
            if section_data and self._has_meaningful_data(section_data):
                data_found = True
                sheet = workbook.create_sheet(sheet_name)
                self._add_template_1_section_data(sheet, data_key, section_data)
                logger.info(f"✅ Created Template 1 sheet: {sheet_name}")
            else:
                logger.info(f"❌ No data found for Template 1 section: {data_key}")
        
        # If no standard sections found, try flexible mapping
        if not data_found:
            logger.info("No standard Template 1 sections found. Attempting flexible mapping...")
            data_found = self._flexible_template_mapping(workbook, data, 1)
        
        return data_found

    def _create_template_2_sheets(self, workbook: Workbook, data: Dict[str, Any]) -> bool:
        """Create sheets for Template 2 - Portfolio Summary Template"""
        logger.info("Creating Template 2 sheets...")
        
        template_config = self.template_configs[2]
        data_found = False
        
        # Process each expected section for Template 2
        for data_key, sheet_name in template_config['sheets'].items():
            section_data = data.get(data_key)
            
            if section_data and self._has_meaningful_data(section_data):
                data_found = True
                sheet = workbook.create_sheet(sheet_name)
                self._add_template_2_section_data(sheet, data_key, section_data)
                logger.info(f"✅ Created Template 2 sheet: {sheet_name}")
            else:
                logger.info(f"❌ No data found for Template 2 section: {data_key}")
        
        # If no standard sections found, try flexible mapping
        if not data_found:
            logger.info("No standard Template 2 sections found. Attempting flexible mapping...")
            data_found = self._flexible_template_mapping(workbook, data, 2)
        
        return data_found

    def _flexible_template_mapping(self, workbook: Workbook, data: Dict[str, Any], template_id: int) -> bool:
        """Flexible mapping of available data to template sections"""
        template_config = self.template_configs[template_id]
        data_found = False
        
        available_keys = [key for key in data.keys() if not key.startswith('_') and self._has_meaningful_data(data.get(key))]
        logger.info(f"Available data keys for mapping: {available_keys}")
        
        # Try to match available keys with template sections
        for available_key in available_keys:
            section_data = data.get(available_key)
            if not section_data:
                continue
                
            # Find the best matching template section
            best_match = self._find_best_template_match(available_key, template_config['sheets'].keys(), template_id)
            
            if best_match:
                sheet_name = template_config['sheets'][best_match]
                sheet = workbook.create_sheet(sheet_name)
                
                if template_id == 1:
                    self._add_template_1_section_data(sheet, best_match, section_data)
                else:
                    self._add_template_2_section_data(sheet, best_match, section_data)
                
                logger.info(f"✅ Mapped '{available_key}' to Template {template_id} sheet: {sheet_name}")
                data_found = True
            else:
                # Create additional sheet for unmapped data
                sheet_name = self._truncate_sheet_name(f"Additional_{available_key}")
                sheet = workbook.create_sheet(sheet_name)
                
                if template_id == 1:
                    self._add_template_1_section_data(sheet, available_key, section_data)
                else:
                    self._add_template_2_section_data(sheet, available_key, section_data)
                
                logger.info(f"✅ Created additional Template {template_id} sheet: {sheet_name}")
                data_found = True
        
        return data_found

    def _find_best_template_match(self, data_key: str, template_sections: List[str], template_id: int) -> str:
        """Find the best matching template section for a data key"""
        data_key_lower = data_key.lower()
        
        # Template-specific matching patterns
        template_patterns = {
            1: {
                'fund': ['fund', 'vehicle', 'investment'],
                'manager': ['manager', 'gp', 'general partner'],
                'financial': ['financial', 'position', 'nav', 'irr'],
                'cashflows': ['cashflow', 'lp', 'investor', 'transaction'],
                'companies': ['company', 'portfolio'],
                'investments': ['investment', 'initial', 'commitment'],
                'valuation': ['valuation', 'value'],
                'financials': ['financial', 'revenue', 'ebitda'],
                'history': ['history', 'transaction'],
                'reference': ['reference', 'country', 'currency']
            },
            2: {
                'executive': ['executive', 'summary', 'overview'],
                'schedule': ['schedule', 'investment'],
                'operations': ['operation', 'income', 'revenue'],
                'cashflows': ['cashflow', 'cash'],
                'pcap': ['pcap', 'capital', 'account'],
                'profile': ['profile', 'company'],
                'financials': ['financial', 'revenue', 'ebitda'],
                'footnotes': ['footnote', 'note', 'disclosure'],
                'reference': ['reference', 'country', 'currency']
            }
        }
        
        patterns = template_patterns.get(template_id, {})
        best_score = 0
        best_match = None
        
        for section in template_sections:
            section_lower = section.lower()
            score = 0
            
            # Check for direct matches
            if data_key_lower == section_lower:
                return section
                
            # Check for partial matches
            if data_key_lower in section_lower or section_lower in data_key_lower:
                score += 3
            
            # Check for pattern matches
            for pattern_key, pattern_terms in patterns.items():
                if pattern_key in section_lower:
                    for term in pattern_terms:
                        if term in data_key_lower:
                            score += 2
                            break
            
            # Check for word overlap
            data_words = set(data_key_lower.split('_'))
            section_words = set(section_lower.split('_'))
            common_words = data_words.intersection(section_words)
            score += len(common_words)
            
            if score > best_score:
                best_score = score
                best_match = section
        
        return best_match if best_score >= 2 else None

    def _add_template_1_section_data(self, sheet, section: str, data: Any):
        """Add data for Template 1 with specialized formatting"""
        row_idx = 1
        
        # Section title with Template 1 color
        title_cell = sheet.cell(row=row_idx, column=1, value=section.replace('_', ' '))
        title_cell.font = Font(size=14, bold=True, color=self.colors['header_blue'])
        row_idx += 2
        
        # Template 1 specific formatting
        if section == 'LP_Investor_Cashflows':
            self._add_cashflows_table(sheet, data, row_idx)
        elif section == 'Fund_Companies':
            self._add_companies_table(sheet, data, row_idx, 'Fund Companies')
        elif section == 'Initial_Investments':
            self._add_investments_table(sheet, data, row_idx, 'Initial Investments')
        elif section == 'Company_Investment_Positions':
            self._add_investment_positions_table(sheet, data, row_idx)
        elif section == 'Company_Valuation':
            self._add_valuation_table(sheet, data, row_idx)
        elif section == 'Company_Financials':
            self._add_financials_table(sheet, data, row_idx)
        elif section == 'Investment_History':
            self._add_investment_history_table(sheet, data, row_idx)
        elif section == 'Reference_Values':
            self._add_reference_values(sheet, data, row_idx)
        else:
            # Default structured data for other sections
            if isinstance(data, dict):
                row_idx = self._add_structured_dict_data(sheet, data, row_idx, section, template_id=1)
            elif isinstance(data, list):
                row_idx = self._add_structured_list_data(sheet, data, row_idx, section, template_id=1)
            else:
                sheet.cell(row=row_idx, column=1, value="Value:").font = Font(bold=True)
                sheet.cell(row=row_idx, column=2, value=self._format_value(data))
        
        self._auto_adjust_columns(sheet)

    def _add_template_2_section_data(self, sheet, section: str, data: Any):
        """Add data for Template 2 with specialized formatting"""
        row_idx = 1
        
        # Section title with Template 2 color
        title_cell = sheet.cell(row=row_idx, column=1, value=section.replace('_', ' '))
        title_cell.font = Font(size=14, bold=True, color=self.colors['accent_green'])
        row_idx += 2
        
        # Template 2 specific formatting
        if section == 'Schedule_of_Investments':
            self._add_schedule_of_investments_table(sheet, data, row_idx)
        elif section == 'Statement_of_Operations':
            self._add_operations_statement(sheet, data, row_idx)
        elif section == 'Statements_of_Cashflows':
            self._add_cashflow_statement(sheet, data, row_idx)
        elif section == 'PCAP_Statements':
            self._add_pcap_statement(sheet, data, row_idx)
        elif section == 'Portfolio_Companies_Profile':
            self._add_companies_table(sheet, data, row_idx, 'Portfolio Companies Profile')
        elif section == 'Portfolio_Companies_Financials':
            self._add_portfolio_financials_table(sheet, data, row_idx)
        elif section == 'FootNotes':
            self._add_footnotes(sheet, data, row_idx)
        elif section == 'Reference_Values':
            self._add_reference_values(sheet, data, row_idx)
        else:
            # Default structured data for other sections
            if isinstance(data, dict):
                row_idx = self._add_structured_dict_data(sheet, data, row_idx, section, template_id=2)
            elif isinstance(data, list):
                row_idx = self._add_structured_list_data(sheet, data, row_idx, section, template_id=2)
            else:
                sheet.cell(row=row_idx, column=1, value="Value:").font = Font(bold=True)
                sheet.cell(row=row_idx, column=2, value=self._format_value(data))
        
        self._auto_adjust_columns(sheet)

    def _add_cashflows_table(self, sheet, data: List, start_row: int):
        """Add LP Investor Cashflows table for Template 1"""
        if not data or not isinstance(data, list):
            sheet.cell(row=start_row, column=1, value="No cashflow data available")
            return
        
        headers = ['Transaction Date', 'Investor Name', 'Transaction Type', 'Amount', 'Currency', 'Description']
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.colors['header_blue'], fill_type="solid")
            cell.border = self._get_border()
            cell.alignment = Alignment(horizontal="center")
        
        current_row = start_row + 1
        
        # Add data rows
        for row_idx, item in enumerate(data):
            if not isinstance(item, dict):
                continue
                
            fill_color = self.colors['light_gray'] if row_idx % 2 == 0 else self.colors['white']
            row_fill = PatternFill(start_color=fill_color, fill_type="solid")
            
            values = [
                item.get('Transaction_Date', ''),
                item.get('Investor_Name', ''),
                item.get('Transaction_Type', ''),
                item.get('Amount'),
                item.get('Currency', ''),
                item.get('Description', '')
            ]
            
            for col_idx, value in enumerate(values, 1):
                cell = sheet.cell(row=current_row, column=col_idx, value=self._format_financial_value(value, headers[col_idx-1]))
                cell.fill = row_fill
                cell.border = self._get_border()
                
                # Format financial columns
                if col_idx == 4 and isinstance(value, (int, float)):
                    cell.number_format = '"$"#,##0'
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in [1, 5]:
                    cell.alignment = Alignment(horizontal="center")
            
            current_row += 1

    def _add_schedule_of_investments_table(self, sheet, data: List, start_row: int):
        """Add Schedule of Investments table for Template 2"""
        if not data or not isinstance(data, list):
            sheet.cell(row=start_row, column=1, value="No investment schedule data available")
            return
        
        headers = [
            'Company Name', 'Fund Name', 'Reported Date', 'Investment Status',
            'Security Type', 'Ownership Percentage', 'Initial Investment Date',
            'Fund Commitment', 'Total Invested', 'Current Cost', 'Reported Value', 'Realized Proceeds'
        ]
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.colors['accent_green'], fill_type="solid")
            cell.border = self._get_border()
            cell.alignment = Alignment(horizontal="center")
        
        current_row = start_row + 1
        
        # Add data rows
        for row_idx, item in enumerate(data):
            if not isinstance(item, dict):
                continue
                
            fill_color = self.colors['light_gray'] if row_idx % 2 == 0 else self.colors['white']
            row_fill = PatternFill(start_color=fill_color, fill_type="solid")
            
            values = [
                item.get('Company_Name', ''),
                item.get('Fund_Name', ''),
                item.get('Reported_Date', ''),
                item.get('Investment_Status', ''),
                item.get('Security_Type', ''),
                item.get('Ownership_Percentage'),
                item.get('Initial_Investment_Date', ''),
                item.get('Fund_Commitment'),
                item.get('Total_Invested'),
                item.get('Current_Cost'),
                item.get('Reported_Value'),
                item.get('Realized_Proceeds')
            ]
            
            for col_idx, value in enumerate(values, 1):
                cell = sheet.cell(row=current_row, column=col_idx, value=self._format_financial_value(value, headers[col_idx-1]))
                cell.fill = row_fill
                cell.border = self._get_border()
                
                # Format financial columns
                if col_idx in [8, 9, 10, 11, 12] and isinstance(value, (int, float)):
                    if abs(value) >= 1000000:
                        cell.number_format = '"$"#,##0.0,,"M"'
                    elif abs(value) >= 1000:
                        cell.number_format = '"$"#,##0,"K"'
                    else:
                        cell.number_format = '"$"#,##0'
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx == 6 and isinstance(value, (int, float)):
                    cell.number_format = '0.00%'
                    cell.alignment = Alignment(horizontal="right")
                elif col_idx in [3, 7]:
                    cell.alignment = Alignment(horizontal="center")
            
            current_row += 1

    def _add_operations_statement(self, sheet, data: Dict, start_row: int):
        """Add Statement of Operations for Template 2"""
        if not data or not isinstance(data, dict):
            sheet.cell(row=start_row, column=1, value="No operations statement data available")
            return
        
        current_row = start_row
        
        # Revenue section
        sheet.cell(row=current_row, column=1, value="REVENUE").font = Font(bold=True)
        current_row += 1
        
        revenue_data = data.get('Revenue', {})
        for key, value in revenue_data.items():
            if value is not None:
                sheet.cell(row=current_row, column=1, value=self._format_field_name(key))
                sheet.cell(row=current_row, column=2, value=self._format_financial_value(value, key))
                current_row += 1
        
        current_row += 1
        
        # Expenses section
        sheet.cell(row=current_row, column=1, value="EXPENSES").font = Font(bold=True)
        current_row += 1
        
        expenses_data = data.get('Expenses', {})
        for key, value in expenses_data.items():
            if value is not None:
                sheet.cell(row=current_row, column=1, value=self._format_field_name(key))
                sheet.cell(row=current_row, column=2, value=self._format_financial_value(value, key))
                current_row += 1
        
        current_row += 1
        
        # Net Results section
        sheet.cell(row=current_row, column=1, value="NET RESULTS").font = Font(bold=True)
        current_row += 1
        
        net_results_data = data.get('Net_Results', {})
        for key, value in net_results_data.items():
            if value is not None:
                sheet.cell(row=current_row, column=1, value=self._format_field_name(key))
                sheet.cell(row=current_row, column=2, value=self._format_financial_value(value, key))
                current_row += 1

    def _add_companies_table(self, sheet, data: List, start_row: int, title: str):
        """Add companies table for both templates"""
        if not data or not isinstance(data, list):
            sheet.cell(row=start_row, column=1, value=f"No {title.lower()} data available")
            return
        
        # Determine headers based on data structure
        if data and isinstance(data[0], dict):
            headers = list(data[0].keys())
        else:
            headers = ['Company Name', 'Industry', 'Headquarters', 'Status']
        
        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row, column=col_idx, value=self._format_field_name(header))
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.colors['header_blue'], fill_type="solid")
            cell.border = self._get_border()
            cell.alignment = Alignment(horizontal="center")
        
        current_row = start_row + 1
        
        # Add data rows
        for row_idx, item in enumerate(data):
            if isinstance(item, dict):
                fill_color = self.colors['light_gray'] if row_idx % 2 == 0 else self.colors['white']
                row_fill = PatternFill(start_color=fill_color, fill_type="solid")
                
                for col_idx, header in enumerate(headers, 1):
                    value = item.get(header, '')
                    cell = sheet.cell(row=current_row, column=col_idx, value=self._format_financial_value(value, header))
                    cell.fill = row_fill
                    cell.border = self._get_border()
            else:
                # Simple list item
                cell = sheet.cell(row=current_row, column=1, value=str(item))
                cell.border = self._get_border()
            
            current_row += 1

    # Additional specialized table methods would go here...
    def _add_investments_table(self, sheet, data: List, start_row: int, title: str):
        """Add investments table - simplified implementation"""
        self._add_companies_table(sheet, data, start_row, title)

    def _add_investment_positions_table(self, sheet, data: List, start_row: int):
        """Add investment positions table - simplified implementation"""
        self._add_companies_table(sheet, data, start_row, "Investment Positions")

    def _add_valuation_table(self, sheet, data: List, start_row: int):
        """Add valuation table - simplified implementation"""
        self._add_companies_table(sheet, data, start_row, "Valuation")

    def _add_financials_table(self, sheet, data: List, start_row: int):
        """Add financials table - simplified implementation"""
        self._add_companies_table(sheet, data, start_row, "Financials")

    def _add_investment_history_table(self, sheet, data: List, start_row: int):
        """Add investment history table - simplified implementation"""
        self._add_companies_table(sheet, data, start_row, "Investment History")

    def _add_portfolio_financials_table(self, sheet, data: List, start_row: int):
        """Add portfolio financials table - simplified implementation"""
        self._add_companies_table(sheet, data, start_row, "Portfolio Financials")

    def _add_cashflow_statement(self, sheet, data: Dict, start_row: int):
        """Add cashflow statement - simplified implementation"""
        self._add_operations_statement(sheet, data, start_row)

    def _add_pcap_statement(self, sheet, data: Dict, start_row: int):
        """Add PCAP statement - simplified implementation"""
        self._add_operations_statement(sheet, data, start_row)

    def _add_footnotes(self, sheet, data: Dict, start_row: int):
        """Add footnotes - simplified implementation"""
        self._add_operations_statement(sheet, data, start_row)

    def _add_reference_values(self, sheet, data: Dict, start_row: int):
        """Add reference values"""
        if not data or not isinstance(data, dict):
            sheet.cell(row=start_row, column=1, value="No reference values available")
            return
        
        current_row = start_row
        
        for category, values in data.items():
            if values and isinstance(values, list):
                sheet.cell(row=current_row, column=1, value=self._format_field_name(category)).font = Font(bold=True)
                current_row += 1
                
                for value in values:
                    if value:
                        sheet.cell(row=current_row, column=2, value=str(value))
                        current_row += 1
                
                current_row += 1

    def _add_structured_dict_data(self, sheet, data: Dict, start_row: int, section: str, template_id: int = 1) -> int:
        """Add structured dictionary data with template-specific formatting"""
        current_row = start_row
        
        if not data:
            sheet.cell(row=current_row, column=1, value=f"No {section} data available")
            return current_row + 1
        
        # Choose color based on template
        header_color = self.colors['header_blue'] if template_id == 1 else self.colors['accent_green']
        
        # Headers
        header_fill = PatternFill(start_color=header_color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        sheet.cell(row=current_row, column=1, value="Field").font = header_font
        sheet.cell(row=current_row, column=1).fill = header_fill
        sheet.cell(row=current_row, column=2, value="Value").font = header_font
        sheet.cell(row=current_row, column=2).fill = header_fill
        
        for col in [1, 2]:
            sheet.cell(row=current_row, column=col).border = self._get_border()
            sheet.cell(row=current_row, column=col).alignment = Alignment(horizontal="center")
        
        current_row += 1
        
        # Data rows with financial formatting
        for idx, (key, value) in enumerate(data.items()):
            if value is None or (isinstance(value, (str, list, dict)) and not value):
                continue
                
            fill_color = self.colors['light_gray'] if idx % 2 == 0 else self.colors['white']
            row_fill = PatternFill(start_color=fill_color, fill_type="solid")
            
            # Field name
            field_cell = sheet.cell(row=current_row, column=1, value=self._format_field_name(key))
            field_cell.fill = row_fill
            field_cell.border = self._get_border()
            
            # Value with financial formatting
            value_cell = sheet.cell(row=current_row, column=2, value=self._format_financial_value(value, key))
            value_cell.fill = row_fill
            value_cell.border = self._get_border()
            
            # Apply number formatting for financial data
            if self._is_financial_field(key) and isinstance(value, (int, float)):
                if abs(value) >= 1000000:
                    value_cell.number_format = '"$"#,##0.0,,"M"'
                elif abs(value) >= 1000:
                    value_cell.number_format = '"$"#,##0,"K"'
                else:
                    value_cell.number_format = '"$"#,##0'
                value_cell.alignment = Alignment(horizontal="right")
            elif isinstance(value, (int, float)):
                value_cell.alignment = Alignment(horizontal="right")
            
            current_row += 1
        
        return current_row

    def _add_structured_list_data(self, sheet, data: List, start_row: int, section: str, template_id: int = 1) -> int:
        """Add structured list data as tables with template-specific formatting"""
        current_row = start_row
        
        if not data:
            sheet.cell(row=current_row, column=1, value=f"No {section} data available")
            return current_row + 1
        
        # Choose color based on template
        header_color = self.colors['header_blue'] if template_id == 1 else self.colors['accent_green']
        
        # If list contains simple values
        if all(not isinstance(item, (dict, list)) or item is None for item in data):
            sheet.cell(row=current_row, column=1, value=section).font = Font(bold=True)
            current_row += 1
            
            for idx, item in enumerate(data):
                sheet.cell(row=current_row, column=1, value=f"{idx + 1}.")
                sheet.cell(row=current_row, column=2, value=self._format_financial_value(item, section))
                current_row += 1
            
            return current_row
        
        # If list contains dictionaries (table data)
        headers = self._get_table_headers(data, section)
        
        if not headers:
            sheet.cell(row=current_row, column=1, value="No valid table data found")
            return current_row + 1
        
        # Table headers
        header_fill = PatternFill(start_color=header_color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=current_row, column=col_idx, value=self._format_field_name(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.border = self._get_border()
            cell.alignment = Alignment(horizontal="center")
        
        current_row += 1
        
        # Table rows with financial formatting
        for row_idx, row_data in enumerate(data):
            if not isinstance(row_data, dict):
                continue
                
            fill_color = self.colors['light_gray'] if row_idx % 2 == 0 else self.colors['white']
            row_fill = PatternFill(start_color=fill_color, fill_type="solid")
            
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header)
                cell = sheet.cell(row=current_row, column=col_idx, value=self._format_financial_value(value, header))
                cell.fill = row_fill
                cell.border = self._get_border()
                
                # Apply financial formatting
                if self._is_financial_field(header) and isinstance(value, (int, float)):
                    if abs(value) >= 1000000:
                        cell.number_format = '"$"#,##0.0,,"M"'
                    elif abs(value) >= 1000:
                        cell.number_format = '"$"#,##0,"K"'
                    else:
                        cell.number_format = '"$"#,##0'
                    cell.alignment = Alignment(horizontal="right")
                elif isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                elif self._is_date_field(header) and self._looks_like_date(value):
                    cell.alignment = Alignment(horizontal="center")
            
            current_row += 1
        
        return current_row

    def _get_table_headers(self, data: List[Dict], section: str) -> List[str]:
        """Get appropriate headers for financial tables"""
        if not data or not isinstance(data[0], dict):
            return []
        
        # Common financial field priorities
        financial_priority_fields = {
            'Company_Name': 100, 'Investment_Date': 95, 'Amount': 90, 'Value': 85,
            'Transaction_Date': 80, 'Investor_Name': 75, 'Currency': 70,
            'Fund_Name': 65, 'NAV': 60, 'IRR': 55, 'Commitment': 50
        }
        
        all_headers = []
        for item in data:
            if isinstance(item, dict):
                for key in item.keys():
                    if key not in all_headers:
                        all_headers.append(key)
        
        # Sort headers by priority
        def get_priority(header):
            return financial_priority_fields.get(header, 0)
        
        return sorted(all_headers, key=get_priority, reverse=True)

    def _format_financial_value(self, value: Any, field_name: str) -> Any:
        """Format values with financial context"""
        if value is None:
            return "N/A"
        elif isinstance(value, bool):
            return "Yes" if value else "No"
        elif isinstance(value, (int, float)):
            # Apply financial formatting based on field name
            if self._is_financial_field(field_name):
                try:
                    if value == 0:
                        return "0"
                    elif abs(value) >= 1000000:
                        return f"${value/1000000:.1f}M"
                    elif abs(value) >= 1000:
                        return f"${value/1000:.0f}K"
                    else:
                        return f"${value:,.0f}"
                except:
                    return f"{value:,.0f}" if value == int(value) else f"{value:,.2f}"
            else:
                return f"{value:,.0f}" if value == int(value) else f"{value:,.2f}"
        elif isinstance(value, list):
            if not value:
                return ""
            return ", ".join(str(self._format_financial_value(item, field_name)) for item in value[:3]) + ("..." if len(value) > 3 else "")
        elif isinstance(value, dict):
            return f"Object({len(value)} fields)"
        elif self._looks_like_date(str(value)):
            return self._format_date_value(str(value))
        else:
            cleaned = str(value).strip()
            return cleaned[:50] + "..." if len(cleaned) > 50 else cleaned

    def _is_financial_field(self, field_name: str) -> bool:
        """Check if field contains financial data"""
        financial_indicators = [
            'amount', 'value', 'price', 'cost', 'nav', 'irr', 'commitment',
            'capital', 'investment', 'fee', 'rate', 'percentage', 'money',
            'fund', 'cash', 'distribution', 'contribution', 'proceeds',
            'revenue', 'income', 'expense', 'ebitda', 'profit', 'loss'
        ]
        field_lower = field_name.lower()
        return any(indicator in field_lower for indicator in financial_indicators)

    def _is_date_field(self, field_name: str) -> bool:
        """Check if field contains date data"""
        date_indicators = ['date', 'time', 'year', 'month', 'day', 'period']
        field_lower = field_name.lower()
        return any(indicator in field_lower for indicator in date_indicators)

    def _looks_like_date(self, value: str) -> bool:
        """Check if string looks like a date"""
        if not isinstance(value, str):
            return False
        date_patterns = [
            r'\d{4}-\d{2}-\d{2}',  # YYYY-MM-DD
            r'\d{2}/\d{2}/\d{4}',   # MM/DD/YYYY
            r'\d{2}-\d{2}-\d{4}',   # MM-DD-YYYY
            r'\d{4}/\d{2}/\d{2}',   # YYYY/MM/DD
        ]
        import re
        return any(re.search(pattern, value) for pattern in date_patterns)

    def _format_date_value(self, value: str) -> str:
        """Format date values consistently"""
        try:
            # Try to parse common date formats
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                try:
                    dt = datetime.strptime(value, fmt)
                    return dt.strftime('%Y-%m-%d')
                except ValueError:
                    continue
            return value
        except:
            return value

    def _create_generic_sheets(self, workbook: Workbook, data: Dict[str, Any]) -> bool:
        """Create generic sheets when template is unknown"""
        logger.info("Creating generic sheets...")
        
        if not data:
            sheet = workbook.create_sheet("No_Data")
            sheet.cell(1, 1, value="❌ No data available for Excel generation")
            return False
        
        data_found = False
        # Create sheets for each top-level key
        for key, value in data.items():
            if key.startswith('_'):
                continue  # Skip metadata
                
            if value and self._has_meaningful_data(value):
                data_found = True
                sheet_name = self._truncate_sheet_name(f"Data_{key}")
                sheet = workbook.create_sheet(sheet_name)
                self._add_structured_dict_data(sheet, {key: value}, 1, key)
                logger.info(f"✅ Created generic sheet: {sheet_name}")
        
        return data_found

    def _format_field_name(self, field_name: str) -> str:
        """Format field names for display"""
        if not field_name:
            return "Field"
        
        formatted = str(field_name)
        formatted = formatted.replace('_', ' ')
        formatted = ' '.join(word.capitalize() for word in formatted.split())
        return formatted

    def _format_value(self, value: Any) -> Any:
        """Format values for Excel display (backward compatibility)"""
        return self._format_financial_value(value, "")

    def _has_meaningful_data(self, data: Any) -> bool:
        """Check if data contains meaningful content"""
        if data is None:
            return False
        elif isinstance(data, (str, int, float, bool)):
            return True
        elif isinstance(data, dict):
            return any(self._has_meaningful_data(v) for v in data.values())
        elif isinstance(data, list):
            return len(data) > 0 and any(self._has_meaningful_data(item) for item in data)
        return False

    def _count_data_points(self, data: Any) -> int:
        """Count non-null, non-empty data points"""
        count = 0
        
        def count_recursive(obj):
            nonlocal count
            if isinstance(obj, dict):
                for v in obj.values():
                    count_recursive(v)
            elif isinstance(obj, list):
                for item in obj:
                    count_recursive(item)
            elif obj is not None and obj != "":
                if isinstance(obj, str) and obj.strip():
                    count += 1
                elif not isinstance(obj, str):
                    count += 1
        
        count_recursive(data)
        return count

    def _get_border(self):
        """Get standard border style"""
        return Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths with financial data considerations"""
        for column in sheet.columns:
            if not column:
                continue
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            
            # Adjust width based on content type
            if any(cell.value and self._is_financial_field(str(cell.value)) for cell in column if cell.value):
                adjusted_width = min(max_length + 4, 20)  # Wider for financial data
            else:
                adjusted_width = min(max_length + 2, 50)
            
            adjusted_width = max(adjusted_width, 8)
            
            try:
                sheet.column_dimensions[column_letter].width = adjusted_width
            except:
                pass

    def _truncate_sheet_name(self, name: str) -> str:
        """Truncate sheet name to 31 characters"""
        return name[:31]