# excel_generator.py
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging
from typing import Dict, Any, List
import os

logger = logging.getLogger(__name__)

class ExcelGenerator:
    def __init__(self):
        self.templates_dir = "templates"
    
    def generate_excel(self, structured_data: Dict[str, Any], template_id: int, output_path: str):
        """
        Generate Excel file from structured data according to template
        """
        try:
            workbook = Workbook()
            
            # Remove default sheet
            if 'Sheet' in workbook.sheetnames:
                del workbook['Sheet']
            
            logger.info(f"Generating Excel with template {template_id}")
            logger.info(f"Data keys: {list(structured_data.keys())}")
            
            # Create sheets based on template
            if template_id == 1:
                self._create_template_1_sheets(workbook, structured_data)
            elif template_id == 2:
                self._create_template_2_sheets(workbook, structured_data)
            else:
                raise ValueError(f"Unknown template ID: {template_id}")
            
            # Save the workbook
            workbook.save(output_path)
            logger.info(f"Excel file successfully generated: {output_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"Error generating Excel file: {str(e)}")
            raise Exception(f"Failed to generate Excel file: {str(e)}")
    
    def _create_template_1_sheets(self, workbook: Workbook, data: Dict[str, Any]):
        """Create sheets for Template 1 - Private Equity Fund Detailed Template"""
        sheets_structure = {
            "Fund and Investment Vehicle Information": {
                "sections": ["Fund Details", "Key Dates", "Fee Structure", "Investment Focus"],
                "description": "Standard details about the Fund, Fund Partnership, and Investment Vehicle"
            },
            "Fund Manager": {
                "sections": ["Management Company", "Firm Details"],
                "description": "Standard details about the Manager (GP) who is running the fund"
            },
            "Fund Investment Vehicle Financial Position": {
                "sections": ["Commitment Summary", "Capital Account", "Performance Metrics", "Carried Interest"],
                "description": "Current cumulative financial position of the Investment Vehicle as of the reporting date"
            },
            "LP Investor cashflows": {
                "sections": ["Cashflow Transactions"],
                "description": "Comprehensive list of transactions between the Investment Vehicle and its investors",
                "is_table": True
            },
            "Fund Companies": {
                "sections": ["Portfolio Companies"],
                "description": "Key details of the companies in which the fund has invested",
                "is_table": True
            },
            "Initial Investments": {
                "sections": ["Investment Details"],
                "description": "Positions of the Investment Vehicle in invested companies as of the investment date",
                "is_table": True
            },
            "Company Investment Positions": {
                "sections": ["Current Positions"],
                "description": "Current positions in invested companies as of the reporting date",
                "is_table": True
            },
            "Company Valuation": {
                "sections": ["Valuation Details"],
                "description": "Valuation details of companies at the current reporting or exit date",
                "is_table": True
            },
            "Company Financials": {
                "sections": ["Financial Information"],
                "description": "Most recent profit & loss, balance sheet, and debt maturity information",
                "is_table": True
            },
            "Investment History": {
                "sections": ["Transaction History"],
                "description": "Full list of historical transactions between the Investment Vehicle and portfolio companies",
                "is_table": True
            },
            "Reference Values": {
                "sections": ["Reference Data"],
                "description": "List of accepted values for dropdown fields"
            }
        }
        
        self._create_all_sheets(workbook, data, sheets_structure)
    
    def _create_template_2_sheets(self, workbook: Workbook, data: Dict[str, Any]):
        """Create sheets for Template 2 - Portfolio Summary Template"""
        sheets_structure = {
            "Executive Portfolio Summary": {
                "sections": ["General Partner", "Portfolio Overview", "Fund Summary", "Performance Metrics", "Portfolio Breakdown"],
                "description": "Executive summary of portfolio performance and composition"
            },
            "Schedule of Investments": {
                "sections": ["Investment Schedule"],
                "description": "Detailed schedule of investments with transparency for Limited Partners",
                "is_table": True
            },
            "Statement of Operations": {
                "sections": ["Revenue", "Expenses", "Net Results"],
                "description": "Fund operations financial statement"
            },
            "Statements of Cashflows": {
                "sections": ["Operating Activities", "Financing Activities", "Net Change"],
                "description": "Cashflow statement showing fund liquidity"
            },
            "PCAP Statements": {
                "sections": ["Beginning Balance", "Cash Flows", "Fees and Expenses", "Investment Activity", "Ending Balance"],
                "description": "Partner Capital Account Statements"
            },
            "Portfolio Companies Profile": {
                "sections": ["Company Profiles"],
                "description": "Detailed profiles of portfolio companies",
                "is_table": True
            },
            "Portfolio Companies Financials": {
                "sections": ["Financial Data"],
                "description": "Financial performance of portfolio companies",
                "is_table": True
            },
            "FootNotes": {
                "sections": ["Fund Organization", "Accounting Policies", "Significant Policies", "Related Party Transactions", "Risk Factors"],
                "description": "Complete and detailed footnote disclosures supporting the financial statements"
            },
            "Reference Values": {
                "sections": ["Reference Data"],
                "description": "Standard reference data for classifications"
            }
        }
        
        self._create_all_sheets(workbook, data, sheets_structure)
    
    def _create_all_sheets(self, workbook: Workbook, data: Dict[str, Any], sheets_config: Dict[str, Any]):
        """Create all sheets with proper data handling"""
        for sheet_name, config in sheets_config.items():
            try:
                sheet_data = data.get(sheet_name, {})
                sections = config.get("sections", [])
                is_table_sheet = config.get("is_table", False)
                description = config.get("description", "")
                
                logger.info(f"Creating sheet '{sheet_name}'")
                
                # Create sheet with truncated name for Excel compatibility
                sheet = workbook.create_sheet(self._truncate_sheet_name(sheet_name))
                
                # Create the structured sheet
                self._create_structured_sheet(sheet, sheet_name, sheet_data, sections, is_table_sheet, description)
                
            except Exception as e:
                logger.error(f"Error creating sheet {sheet_name}: {e}")
                # Create empty sheet with error message
                try:
                    sheet = workbook.create_sheet(self._truncate_sheet_name(sheet_name))
                    sheet.cell(1, 1, value=f"Error creating sheet: {str(e)}")
                except:
                    pass
    
    def _truncate_sheet_name(self, name: str) -> str:
        """Truncate sheet name to 31 characters for Excel compatibility"""
        return name[:31]
    
    def _create_structured_sheet(self, sheet, sheet_name: str, data: Any, sections: List[str], is_table_sheet: bool = False, description: str = ""):
        """Create a structured sheet with data"""
        row_idx = 1
        
        # Sheet title
        title_cell = sheet.cell(row=row_idx, column=1, value=sheet_name)
        title_cell.font = Font(size=16, bold=True, color="2E86AB")
        row_idx += 1
        
        # Sheet description
        if description:
            desc_cell = sheet.cell(row=row_idx, column=1, value=description)
            desc_cell.font = Font(size=10, italic=True, color="666666")
            row_idx += 1
        
        row_idx += 1  # Add spacing
        
        # If data is empty or contains error, show message
        if not data or (isinstance(data, dict) and data.get("error")):
            error_msg = data.get("error", "No data available for this section") if isinstance(data, dict) else "No data available for this section"
            sheet.cell(row=row_idx, column=1, value=f"Note: {error_msg}").font = Font(color="FF0000", italic=True)
            row_idx += 2
        
        # Handle table-style sheets (where the main data is a list)
        if is_table_sheet and isinstance(data, list) and data:
            logger.info(f"Processing table data for {sheet_name} with {len(data)} rows")
            row_idx = self._add_table_section(sheet, data, row_idx, f"{sheet_name} Data")
        elif is_table_sheet and isinstance(data, dict) and sections:
            # Try to find list data in the sections
            for section in sections:
                section_data = data.get(section)
                if isinstance(section_data, list) and section_data:
                    logger.info(f"Found table data in section '{section}' with {len(section_data)} rows")
                    row_idx = self._add_table_section(sheet, section_data, row_idx, section)
                    row_idx += 2
        else:
            # Process regular sections
            for section in sections:
                # Section header
                sheet.cell(row=row_idx, column=1, value=section).font = Font(size=14, bold=True, color="2E86AB")
                row_idx += 1
                
                # Get section data
                section_data = self._extract_section_data(data, section)
                
                if not section_data or (isinstance(section_data, dict) and not section_data):
                    sheet.cell(row=row_idx, column=1, value="No data available for this section").font = Font(italic=True, color="888888")
                    row_idx += 2
                    continue
                
                if isinstance(section_data, dict):
                    row_idx = self._add_key_value_section(sheet, section_data, row_idx)
                    row_idx += 2  # Spacing between sections
                    
                elif isinstance(section_data, list) and section_data:
                    row_idx = self._add_table_section(sheet, section_data, row_idx, section)
                    row_idx += 2  # Spacing
                    
                else:
                    # Single value
                    sheet.cell(row=row_idx, column=1, value="Value").font = Font(bold=True)
                    sheet.cell(row=row_idx, column=2, value=self._format_value(section_data))
                    row_idx += 3
        
        # Auto-adjust columns
        self._auto_adjust_columns(sheet)
        
        # Add summary note if no data was added
        if row_idx <= 5:  # Only title and minimal content
            sheet.cell(row=row_idx, column=1, value="No structured data extracted from document").font = Font(italic=True, color="FF6600")
    
    def _extract_section_data(self, data: Any, section: str) -> Any:
        """Extract data for a specific section with flexible handling"""
        if isinstance(data, dict):
            # Try direct key access first
            if section in data:
                return data[section]
            
            # Try case-insensitive search
            section_lower = section.lower()
            for key, value in data.items():
                if key.lower() == section_lower:
                    return value
            
            # Try partial match
            for key, value in data.items():
                if section_lower in key.lower() or key.lower() in section_lower:
                    return value
            
            return {}
        elif isinstance(data, list):
            return data
        else:
            return data
    
    def _add_key_value_section(self, sheet, data: Dict, start_row: int) -> int:
        """Add key-value pairs to sheet with enhanced formatting"""
        current_row = start_row
        
        # Headers
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        sheet.cell(row=current_row, column=1, value="Field").font = header_font
        sheet.cell(row=current_row, column=1).fill = header_fill
        sheet.cell(row=current_row, column=2, value="Value").font = header_font
        sheet.cell(row=current_row, column=2).fill = header_fill
        
        # Style headers with borders
        for col in [1, 2]:
            cell = sheet.cell(row=current_row, column=col)
            cell.border = self._get_border()
            cell.alignment = Alignment(horizontal="center")
        
        current_row += 1
        
        # Data rows
        row_fill_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        row_fill_dark = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        row_count = 0
        for field, value in data.items():
            if value is not None and value != "":
                # Alternate row colors
                fill = row_fill_light if row_count % 2 == 0 else row_fill_dark
                
                sheet.cell(row=current_row, column=1, value=self._format_field_name(field))
                sheet.cell(row=current_row, column=1).fill = fill
                sheet.cell(row=current_row, column=2, value=self._format_value(value))
                sheet.cell(row=current_row, column=2).fill = fill
                
                # Add borders
                for col in [1, 2]:
                    sheet.cell(row=current_row, column=col).border = self._get_border()
                
                current_row += 1
                row_count += 1
        
        return current_row
    
    def _add_table_section(self, sheet, data: List[Dict], start_row: int, section_name: str = "") -> int:
        """Add tabular data to sheet with enhanced formatting"""
        if not data:
            sheet.cell(row=start_row, column=1, value="No table data available").font = Font(italic=True, color="888888")
            return start_row + 2
        
        current_row = start_row
        
        # Section header for table
        if section_name:
            sheet.cell(row=current_row, column=1, value=section_name).font = Font(size=12, bold=True, color="2E86AB")
            current_row += 1
        
        # Get all unique headers from all rows
        headers = []
        for row in data:
            if isinstance(row, dict):
                for key in row.keys():
                    if key not in headers:
                        headers.append(key)
        
        if not headers:
            sheet.cell(row=current_row, column=1, value="No valid table headers found").font = Font(italic=True, color="888888")
            return current_row + 2
        
        logger.info(f"Creating table with {len(headers)} headers and {len(data)} rows")
        
        # Table headers
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=current_row, column=col_idx, value=self._format_field_name(header))
            cell.font = header_font
            cell.fill = header_fill
            cell.border = self._get_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        current_row += 1
        
        # Table data rows
        row_fill_light = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        row_fill_dark = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        for row_idx, row_data in enumerate(data):
            if isinstance(row_data, dict):
                # Alternate row colors
                fill = row_fill_light if row_idx % 2 == 0 else row_fill_dark
                
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header)
                    cell = sheet.cell(row=current_row, column=col_idx, value=self._format_value(value))
                    cell.fill = fill
                    cell.border = self._get_border()
                    # Right-align numbers
                    if isinstance(value, (int, float)):
                        cell.alignment = Alignment(horizontal="right")
                
                current_row += 1
        
        logger.info(f"Table created successfully with {len(data)} rows")
        return current_row
    
    def _format_field_name(self, field_name: str) -> str:
        """Format field names for better display"""
        if not field_name:
            return "Field"
        
        # Convert camelCase to Space Separated
        import re
        formatted = re.sub(r'([a-z])([A-Z])', r'\1 \2', str(field_name))
        formatted = re.sub(r'([A-Z])([A-Z][a-z])', r'\1 \2', formatted)
        
        # Capitalize first letter of each word
        formatted = ' '.join(word.capitalize() for word in formatted.split())
        
        return formatted
    
    def _format_value(self, value: Any) -> Any:
        """Format values for Excel display with enhanced handling"""
        if value is None:
            return ""
        elif isinstance(value, (int, float)):
            # Format numbers with commas
            try:
                if value >= 1000 or value <= -1000:
                    return f"{value:,.2f}"
                else:
                    return f"{value:.2f}"
            except:
                return value
        elif isinstance(value, bool):
            return "Yes" if value else "No"
        elif isinstance(value, list):
            if not value:
                return ""
            # Handle list of strings/numbers
            if all(isinstance(item, (str, int, float)) for item in value):
                return ", ".join(str(item) for item in value)
            else:
                return f"Array with {len(value)} items"
        elif isinstance(value, dict):
            # For nested dictionaries, show a summary
            return "Nested data structure"
        else:
            # String values - clean up any extra whitespace
            cleaned = str(value).strip()
            if len(cleaned) > 100:  # Truncate very long values
                return cleaned[:100] + "..."
            return cleaned
    
    def _get_border(self):
        """Get standard border style"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        return thin_border
    
    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths with better handling"""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        # Calculate length considering formatting
                        cell_value = str(cell.value)
                        length = len(cell_value)
                        
                        # Adjust for number formatting
                        if isinstance(cell.value, (int, float)):
                            length += 2  # Add space for formatting
                        
                        if length > max_length:
                            max_length = length
                except:
                    pass
            
            # Set reasonable column width
            adjusted_width = min(max_length + 2, 50)
            adjusted_width = max(adjusted_width, 8)  # Minimum width
            sheet.column_dimensions[column_letter].width = adjusted_width